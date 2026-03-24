import openpyxl
import json
import sys

def dump_formulas(filename):
    wb = openpyxl.load_workbook(filename, data_only=False)
    wb_data = openpyxl.load_workbook(filename, data_only=True)
    
    output = []
    
    for sheet_name in ["BOLSA PLANA MANIJA", "BOLSA PLANA TROQUEL"]:
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]
        
        output.append(f"--- {sheet_name} ---")
        for row in range(1, 60):
            for col in range(1, 15): 
                cell = ws.cell(row=row, column=col)
                cell_data = ws_data.cell(row=row, column=col)
                if cell.value is not None:
                    val_str = str(cell.value)
                    data_str = str(cell_data.value)
                    coord = cell.coordinate
                    output.append(f"  {coord}: {val_str} = {data_str}")
    
    with open("excel_formulas_marzo.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(output))
    print("Saved to excel_formulas_marzo.txt")

if __name__ == "__main__":
    dump_formulas("COSTOS HOJA MADRE - marzo.xlsx")
