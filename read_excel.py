import pandas as pd
import openpyxl
import json

file_path = "COSTOS HOJA MADRE.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    result = {"sheets": wb.sheetnames, "data": {}}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]
        
        sheet_info = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=80, min_col=1, max_col=15), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is not None:
                    try:
                        val = ws_data.cell(row=row_idx, column=col_idx).value
                    except:
                        val = None
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    
                    cell_info = {
                        "cell": f"{col_letter}{row_idx}",
                        "formula_or_value": str(cell.value),
                        "evaluated_value": str(val)
                    }
                    sheet_info.append(cell_info)
        result["data"][sheet_name] = sheet_info

    with open("excel_data.json", "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    print("Parsed successfully. Wrote to excel_data.json.")
        
except Exception as e:
    print(f"Error: {e}")
